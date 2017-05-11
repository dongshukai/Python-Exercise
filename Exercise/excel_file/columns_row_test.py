# -*- coding: utf-8 -*-

# def int_toletter(column):
#         """
#         More information please visit http://support.microsoft.com/kb/833402
#         """
#         result = ""
#         alpha = column // 26
#         # remainder = column - (alpha * 26)
#         remainder = column % 26
#         if alpha > 0:
#             result = chr(alpha + 64)
#         if remainder > 0:
#             result += chr(remainder + 64)
#         if remainder == 0:
#             result += 'Z'
#         return result
#
# # a = int_toletter(104)
# b = int_toletter(52)
# print b

# 65 - A  90-Z
# def my_fun(column):
#     result = ''
#     alpha = column
#     remainder = -1
#     while alpha > 27:
#         remainder = alpha % 26
#         if remainder == 0:
#             result = 'Z' + result
#         else:
#             result = chr(remainder + 64) + result
#         alpha = alpha // 26
#         if remainder == 0:
#             alpha -= 1
#     else:
#         if column > 26:
#             if alpha == 27:
#                 result = 'AA' + result
#             else:
#                 result = chr(alpha + 64) + result
#         else:
#             result = chr(alpha + 64) + result
#     return result
#
# a = my_fun(1)
# z = my_fun(26)
# aa = my_fun(27)
# az = my_fun(52)
# ba = my_fun(53)
# bz = my_fun(78)
# zy = my_fun(701)
# zz = my_fun(702)
# aaa = my_fun(703)
# alz = my_fun(1014)
# ama = my_fun(1015)
# azz = my_fun(1378)
# baa = my_fun(1379)
# kzz = my_fun(8138)
# laa = my_fun(8139)
# x = my_fun(9247)
#
# print 'a: %s, z: %s' % (a, z)
# print "AA: %s, AZ: %s, BA: %s, BZ: %s" % (aa, az, ba, bz)
# print 'ZY: %s, ZZ: %s, AAA: %s' % (zy, zz, aaa)
# print 'AZZ: %s, BAA: %s, KZZ: %s, LAA: %s' % (azz, baa, kzz, laa)
# print x





# import xlwt
# from datetime import datetime
#
# style0 = xlwt.easyxf('font: name Times New Roman, color-index red, bold on',num_format_str='#,##0.00')
# style1 = xlwt.easyxf(num_format_str='D-MMM-YY')
#
# wb = xlwt.Workbook()
# ws = wb.add_sheet('A Test Sheet')
#
# ws.write(0, 0, 1234.56, style0)
# ws.write(1, 0, datetime.now(), style1)
# ws.write(2, 0, 4)
# ws.write(2, 1, 1)
# ws.write(2, 2, xlwt.Formula("SUM(A3,B5)"))
#
# # ws.write(3, 0, 0)
# # ws.write(3, 0, 1)
# # ws.write(3, 0, 2)
# # ws.write(3, 0, 4)
#
#
# wb.save('example.xls')

import openpyxl
from openpyxl.styles import Font, colors, fills, Fill, PatternFill, GradientFill


# from openpyxl.styles import numbers
from cStringIO import StringIO

workbook = openpyxl.Workbook()
worksheet = workbook.active
# excel_data = [[201607, 127],
#               [u'\u6279\u6b21', u'\u5458\u5de5\u7f16\u53f7', u'\u5458\u5de5', u'\u57fa\u672c\u5de5\u8d44', u'\u5b9e\u9645\u57fa\u672c\u5de5\u8d44', u'\u6bcf\u6708\u8ba1\u85aa\u5929\u6570', u'\u7a0e\u91d1\u8c03\u6574', u'\u7a0e\u524d\u53d1\u653e\u8c03\u6574', u'\u8ba1\u7a0e\u6b63\u7b97\u4e0d\u53d1\u7a0e\u524d', u'\u7a0e\u524d\u4e0d\u53d1\u8c03\u6574', u'\u8ba1\u7a0e\u5012\u7b97\u4e0d\u53d1\u7a0e\u540e', u'\u5012\u7b97\u7a0e\u540e\u4e0d\u53d1\u8c03\u6574', u'\u7a0e\u540e\u53d1\u653e\u8c03\u6574', u'\u8ba1\u7a0e\u5012\u7b97\u53d1\u653e\u7a0e\u540e', u'\u7a0e\u540e\u4e0d\u53d1\u8c03\u6574 ', u'\u5012\u7b97\u7a0e\u540e\u53d1\u653e\u8c03\u6574', u'\u5012\u7b97\u7a0e\u540e', u'\u8ba1\u7a0e\u6b63\u7b97\u53d1\u653e\u7a0e\u524d', u'\u6b63\u7b97\u7a0e\u524d', u'\u6b63\u7b97\u7a0e\u91d1', u'\u6b63\u7b97\u7a0e\u540e', u'\u603b\u7a0e\u540e', u'\u603b\u7a0e\u524d', u'\u5012\u7b97\u7a0e\u524d', u'\u603b\u7a0e\u91d1', u'\u5012\u7b97\u7a0e\u91d1', u'\u623f\u8d34', u'\u5b9e\u9645\u623f\u8d34', u'\u9910\u8d34 ', u'\u5b9e\u9645\u9910\u8d34', u'\u4ea4\u901a\u6d25\u8d34', u'\u5b9e\u9645\u4ea4\u901a\u6d25\u8d34', u'\u8f66\u8d34', u'\u5b9e\u9645\u8f66\u8d34', u'\u72ec\u751f\u5b50\u5973\u6d25\u8d34', u'\u5e73\u65f6\u52a0\u73ed\u5c0f\u65f6\u6570', u'\u5468\u672b\u52a0\u73ed\u5c0f\u65f6\u6570', u'\u8282\u5047\u65e5\u52a0\u73ed\u5c0f\u65f6\u6570', u'\u5e73\u65f6\u52a0\u73ed\u6d25\u8d34', u'\u5468\u672b\u52a0\u73ed\u6d25\u8d34', u'\u5047\u65e5\u52a0\u73ed\u6d25\u8d34', u'\u77ed\u75c5\u5047\u5929\u6570', u'\u65f7\u5de5\u5929\u6570', u'\u77ed\u75c5\u5047\u6263\u6b3e', u'\u65f7\u5de5\u6263\u6b3e', u'\u4ea7\u5047\u5de5\u8d44 ', u'\u957f\u75c5\u5047\u5de5\u8d44 ', u'\u6cd5\u5b9a\u5e74\u5047\u4f59\u989d', u'\u516c\u53f8\u5e74\u5047\u4f59\u989d', u'\u6cd5\u5b9a\u5e74\u5047\u6298\u73b0', u'\u516c\u53f8\u5e74\u5047\u6298\u73b0', u'\u4f01\u4e1a\u517b\u8001\u6c47\u7f34\u989d', u'\u4f01\u4e1a\u533b\u7597\u6c47\u7f34\u989d', u'\u4f01\u4e1a\u5931\u4e1a\u6c47\u7f34\u989d', u'\u4f01\u4e1a\u5927\u75c5\u533b\u7597\u6c47\u7f34\u989d', u'\u5de5\u4f24\u4fdd\u9669\u6c47\u7f34\u989d', u'\u751f\u80b2\u4fdd\u9669\u6c47\u7f34\u989d', u'\u6b8b\u4fdd\u91d1\u6c47\u7f34\u989d', u'\u91c7\u6696\u8d39\u6c47\u7f34\u989d', u'\u5de5\u4f1a\u8d39\u6c47\u7f34\u989d', u'\u4f01\u4e1a\u516c\u79ef\u91d1\u6c47\u7f34\u989d', u'\u4f01\u4e1a\u8865\u5145\u516c\u79ef\u91d1\u6c47\u7f34\u989d', u'\u4e2a\u4eba\u517b\u8001\u6c47\u7f34\u989d', u'\u4e2a\u4eba\u533b\u7597\u6c47\u7f34\u989d', u'\u4e2a\u4eba\u5931\u4e1a\u6c47\u7f34\u989d', u'\u4e2a\u4eba\u5927\u75c5\u533b\u7597\u6c47\u7f34\u989d', u'\u4e2a\u4eba\u516c\u79ef\u91d1\u6c47\u7f34\u989d', u'\u4e2a\u4eba\u8865\u5145\u516c\u79ef\u91d1\u6c47\u7f34\u989d', u'\u4f01\u4e1a\u517b\u8001\u8865\u7f34\u989d', u'\u4f01\u4e1a\u533b\u7597\u8865\u7f34\u989d', u'\u4f01\u4e1a\u5931\u4e1a\u8865\u7f34\u989d', u'\u4f01\u4e1a\u5927\u75c5\u533b\u7597\u8865\u7f34\u989d', u'\u5de5\u4f24\u4fdd\u9669\u8865\u7f34\u989d', u'\u751f\u80b2\u4fdd\u9669\u8865\u7f34\u989d', u'\u6b8b\u4fdd\u91d1\u8865\u7f34\u989d', u'\u91c7\u6696\u8d39\u8865\u7f34\u989d', u'\u5de5\u4f1a\u8d39\u8865\u7f34\u989d', u'\u4f01\u4e1a\u516c\u79ef\u91d1\u8865\u7f34\u989d', u'\u4f01\u4e1a\u8865\u5145\u516c\u79ef\u91d1\u8865\u7f34\u989d', u'\u4e2a\u4eba\u517b\u8001\u8865\u7f34\u989d', u'\u4e2a\u4eba\u533b\u7597\u8865\u7f34\u989d', u'\u4e2a\u4eba\u5931\u4e1a\u8865\u7f34\u989d', u'\u4e2a\u4eba\u5927\u75c5\u533b\u7597\u8865\u7f34\u989d', u'\u4e2a\u4eba\u516c\u79ef\u91d1\u8865\u7f34\u989d', u'\u4e2a\u4eba\u793e\u4fdd\u5408\u8ba1', u'\u4f01\u4e1a\u793e\u4fdd\u5408\u8ba1', u'\u4e2a\u4eba\u516c\u79ef\u91d1\u5408\u8ba1', u'\u4f01\u4e1a\u516c\u79ef\u91d1\u5408\u8ba1', u'\u5e94\u53d1\u5de5\u8d44', u'\u5e94\u7a0e\u5de5\u8d44', u'\u516c\u53f8\u627f\u62c5\u5e94\u7a0e', u'\u5458\u5de5\u627f\u62c5\u4e2a\u4eba\u6240\u5f97\u7a0e', u'\u5b9e\u9645\u5e94\u7a0e', u'\u516c\u53f8\u627f\u62c5\u4e2a\u4eba\u6240\u5f97\u7a0e', u'\u7d2f\u79ef\u5e94\u7a0e', u'\u5e74\u7ec8\u5956', u'\u516c\u53f8\u627f\u62c5\u5e74\u7ec8\u5956\u5e94\u7a0e', u'\u5e74\u7ec8\u5956\u5b9e\u9645\u5e94\u7a0e', u'\u5458\u5de5\u627f\u62c5\u5e74\u7ec8\u5956\u7a0e', u'\u516c\u53f8\u627f\u62c5\u5e74\u7ec8\u5956\u7a0e', u'\u5e74\u7ec8\u5956\u7d2f\u79ef\u5e94\u7a0e', u'\u52b3\u52a1\u62a5\u916c', u'\u516c\u53f8\u627f\u62c5\u52b3\u52a1\u62a5\u916c\u5e94\u7a0e', u'\u52b3\u52a1\u62a5\u916c\u5b9e\u9645\u5e94\u7a0e', u'\u5458\u5de5\u627f\u62c5\u52b3\u52a1\u6240\u5f97\u7a0e', u'\u4f01\u4e1a\u627f\u62c5\u52b3\u52a1\u6240\u5f97\u7a0e', u'\u79bb\u804c\u8865\u507f\u91d1', u'\u516c\u53f8\u627f\u62c5\u79bb\u804c\u8865\u507f\u91d1\u5e94\u7a0e', u'\u79bb\u804c\u8865\u507f\u91d1\u5b9e\u9645\u5e94\u7a0e', u'\u4e2a\u4eba\u627f\u62c5\u7684\u79bb\u804c\u8865\u507f\u91d1\u6240\u5f97\u7a0e', u'\u516c\u53f8\u627f\u62c5\u7684\u79bb\u804c\u8865\u507f\u91d1\u6240\u5f97\u7a0e', u'\u80a1\u6743\u6fc0\u52b1\u6240\u5f97', u'\u516c\u53f8\u627f\u62c5\u80a1\u6743\u5e94\u7a0e', u'\u89c4\u5b9a\u6708\u6570', u'\u80a1\u6743\u5b9e\u9645\u5e94\u7a0e', u'\u6388\u6743\u65e5\u80a1\u7968\u4ef7\u683c', u'\u884c\u6743\u65e5\u80a1\u7968\u4ef7\u683c', u'\u4e2a\u4eba\u627f\u62c5\u80a1\u6743\u7a0e', u'\u516c\u53f8\u627f\u62c5\u80a1\u6743\u7a0e', u'\u5076\u7136\u6240\u5f97', u'\u516c\u53f8\u627f\u62c5\u5076\u7136\u6240\u5f97\u5e94\u7a0e', u'\u5076\u7136\u6240\u5f97\u5b9e\u9645\u5e94\u7a0e', u'\u5458\u5de5\u627f\u62c5\u5076\u7136\u6240\u5f97\u7a0e', u'\u516c\u53f8\u627f\u62c5\u5076\u7136\u6240\u5f97\u7a0e', u'\u5458\u5de5\u627f\u62c5\u7684\u4e2a\u7a0e\u5408\u8ba1', u'\u516c\u53f8\u627f\u62c5\u7684\u4e2a\u7a0e\u5408\u8ba1', u'\u5f53\u524d\u6279\u6b21\u7a0e\u524d', u'\u5b9e\u53d1\u5de5\u8d44', u'\u5f53\u524d\u6279\u6b21\u7a0e\u91d1', u'\u5f53\u524d\u6279\u6b21\u7a0e\u540e'], [u'T001-20160701', u'EM20160727002', u'\u7ba1\u7406\u5458', 0.0, 0.0, 21.75, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0], [u'T001-20160701', u'EM20160727003', u'\u914d\u7f6e\u5458', 0.0, 0.0, 21.75, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0], [u'T001-20160701', u'EM20160727004', u'\u64cd\u4f5c\u5458', 0.0, 0.0, 21.75, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0], [u'T001-20160701', u'EM20160727005', u'\u5ba1\u6838\u5458', 0.0, 0.0, 21.75, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
#               ]
# for row_index, row in enumerate(excel_data):
#     for cell_index, cell_value in enumerate(row):
#         worksheet.cell(row=row_index+1, column=cell_index+1, value=cell_value)
# print openpyxl.styles.numbers.BUILTIN_FORMATS
# openpyxl.styles.numbers.BUILTIN_FORMATS[2]

worksheet.cell(row=1, column=1, value=12.0005).number_format = '0.000'
worksheet.cell(row=1, column=2, value=u'明天').number_format = 'General'
cell = worksheet.cell(row=1, column=3, value=u'哈哈') # .fill = fills.Color(rgb='00FF0000')
# cell.fill.fill_type = Fill.FILL_SOLID
# cell.fill.bgColor = colors.YELLOW
# cell.fill.fgColor = colors.GREEN
# cell.fill.end_color.rgb = 'FFFFFFFF'
# cell.fill = PatternFill(bgColor=colors.BLUE, fgColor=colors.RED)
# cell.fill = GradientFill(stop=("00CFFFCE", "00CFFFCE"))
cell.fill = GradientFill(stop=("00FFFF00", "00FFFF00"))

# .style.fill.start_color = colors.Color(rgb='00FF00')


# fp = StringIO()
workbook.save('123.xlsx')
# fp.seek(0)
# excel_file = fp.read()
# fp.close()