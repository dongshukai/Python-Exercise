# -*- coding: utf-8 -*-

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

wb = Workbook()
ws = wb.active
my_cell = ws['B2']
my_cell.value = "My Cell"
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
# side2 = Side(style='medium', color='FF000000')

border = Border(top=double, left=thin, right=thin, bottom=double)
fill = PatternFill("solid", fgColor="DDDDDD")
fill = GradientFill(stop=("000000", "FFFFFF"))
font = Font(b=True, color="FF0000")
al = Alignment(horizontal="center", vertical="center")

ws['D1'] = 'abc'
ws['D2'] = 123
ws['D3'] = '=D1-D2'

ws['E5'] = '一路软件'
ws['E5'].alignment = Alignment(horizontal="center", vertical="center", justifyLastLine=1, wrap_text=False)

ws['F5'] = '易路软件'
ws['F5'].alignment = Alignment(horizontal="center", vertical="center", text_rotation=180, wrap_text=False)

ws['G5'] = '易路软件123'
ws['G5'].alignment = Alignment(horizontal="center", vertical="top", wrap_text=False)

ws.Rows(5).Delete()

wb.save("styled.xlsx")
