# -*- coding: utf-8 -*-
# ******************************************************
__author__ = 'dongsk'
__create_date__ = '16/12/29'
__filename__ = 'excel_merge_test.py'
# ******************************************************

from openpyxl import Workbook
from openpyxl.styles import Font, colors, GradientFill, Alignment, Border, Side, PatternFill


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

wb = Workbook()
ws = wb.active
ws.title = 'Expense Summary'

ws['A1'] = 'Asia Infrastructure Investment Bank'
ws['A1'].font = Font(name='Arial', size=20, bold=True, italic=False, vertAlign=None,
                     underline='none', strike=False, color='FF000000')

ws['A2'] = 'Expense Summary'
ws['A2'].font = Font(name='Arial', size=20, bold=True, italic=False, vertAlign=None,
                     underline='none', strike=False, color='FF000000')

ws['A3'] = 'Period'
ws['A4'] = 'Create by'
ws['A5'] = 'Create on'

ws['A10'] = 'Expense Report ID'
ws['A10'].fill = GradientFill(stop=('00C0C0C0', '00C0C0C0'))
ws['A10'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws['F10'] = 'Applicant'
ws['F10'].fill = GradientFill(stop=('0000FF7F', '0000FF7F'))
ws['F10'].alignment = Alignment(horizontal="center", vertical="center")


ws.row_dimensions[8].height = 0.33 * 72
ws.row_dimensions[9].height = 0.39 * 72
ws.row_dimensions[10].height = 0.54 * 72

ws.row_dimensions[10].font = Font(name='Arial', size=10, bold=True, italic=False, vertAlign=None,
                                  underline='none', strike=False, color='FF000000')

ws.merge_cells('F8:L9')

ws['F8'] = 'Employee Information'
ws['F8'].fill = GradientFill(stop=('0000FF7F', '0000FF7F'))
ws['F8'].alignment = Alignment(horizontal="center", vertical="center")


ws.merge_cells('AT8:AT9')

# ws.merge_cells('A1:B1')
#
# ws.merge_cells('A3:C6')
#
# ws.merge_cells('D2:D5')


wb.save("styled.xlsx")
